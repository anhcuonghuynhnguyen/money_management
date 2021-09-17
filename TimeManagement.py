import time
from tkinter import *
from tkinter import ttk 
from time import strftime
from PIL import ImageTk
from openpyxl import *

wb = Workbook()
ws = wb.active

"------------------------------------------------------------------------------------------------------------------------------------------------"

# define fuction for Thu vào

def update_thu():
    # update Stt
    row = ws.max_row

    Stt_entry.delete(0, END)
    Stt_entry.insert(END, row)

    # update month
    Month_entry.delete(0, END)
    Month_entry.insert(END, time.strftime("%m/%Y"))

    # update bama cho
    Bama_entry.delete(0, END)
    Bama_entry.insert(END, "3000000")

    # update lương
    Luong_entry.delete(0, END)
    Luong_entry.insert(END, "2000000")

def delete_thu():
    # delete all
    Stt_entry.delete(0, END)
    Month_entry.delete(0, END)
    Bama_entry.delete(0, END)
    Luong_entry.delete(0, END)
    Ghi_chu.delete("1.0", END)

"------------------------------------------------------------------------------------------------------------------------------------------------"

# define function for Chi ra

def update_chi():
    delete_chi()
    an_entry.insert(END, "1000000")
    nha_entry.insert(END, "1000000")
    dien_entry.insert(END, "50000")
    nuoc_entry.insert(END, "10000")
    xang_entry.insert(END, "200000")
    vdsh_entry.insert(END, "100000")
    toeic_entry.insert(END, "300000")
    khac_entry.insert(END, "300000")
    dt_entry.insert(END, "50000")
    thuoc_entry.insert(END, "50000")
    pass

def delete_chi( a = "0"):
    if a == "1":
        an_entry.delete(0, END)
        nha_entry.delete(0, END)
        dien_entry.delete(0, END)
        nuoc_entry.delete(0, END)
        xang_entry.delete(0, END)
        dt_entry.delete(0, END)
        vdsh_entry.delete(0, END)
        toeic_entry.delete(0, END)
        toeic_entry.delete(0, END)
        khac_entry.delete(0, END)
        thuoc_entry.delete(0, END)
        Ghi_chu1.delete("1.0", END)
        
    else:
        an_entry.delete(0, END)
        nha_entry.delete(0, END)
        dien_entry.delete(0, END)
        nuoc_entry.delete(0, END)
        xang_entry.delete(0, END)
        dt_entry.delete(0, END)
        vdsh_entry.delete(0, END)
        toeic_entry.delete(0, END)
        khac_entry.delete(0, END)
        thuoc_entry.delete(0, END)


"------------------------------------------------------------------------------------------------------------------------------------------------"
# define fuction for Báo cáo

def updateBaoCao():
    Tongthu()
    Tongchi()
    Chichiem()
    Tietkiem()
    

def Tongthu():
    global Sum
    Sum = int(Bama_entry.get()) + int(Luong_entry.get())
    thu_entry.configure(state="normal")
    thu_entry.delete(0, END)
    thu_entry.insert(END, Sum)
    thu_entry.configure(state= "readonly")

def Tongchi():
    global Chi
    Chi = (int(an_entry.get()) + int(nha_entry.get()) + int(dien_entry.get()) 
                + int(nuoc_entry.get()) + int(xang_entry.get()) + int(dt_entry.get()) 
                + int(vdsh_entry.get()) + int(toeic_entry.get()) + int(thuoc_entry.get())
                + int(khac_entry.get())
    )
    chi_entry.configure(state= "normal")
    chi_entry.delete(0, END)
    chi_entry.insert(END, Chi)
    chi_chiem_entry.configure(state= "readonly")


def Chichiem():
    Ti_le = round((int(Chi)/ int(Sum)) * 100, 2)
    chi_chiem_entry.configure(state= "normal")
    chi_chiem_entry.delete(0, END)
    chi_chiem_entry.insert(END, str(Ti_le) + "%")
    chi_chiem_entry.configure(state= "readonly")

def Tietkiem():
    Ti_kiem = int(Sum) - int(Chi)
    tiet_kiem_entry.configure(state= "normal")
    tiet_kiem_entry.delete(0, END)
    tiet_kiem_entry.insert(END, Ti_kiem)
    tiet_kiem_entry.configure(state= "readonly")

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set function for button save
def Save():
    row = int(ws.max_row) + 1
    ws["A"+str(row)].value = Stt_entry.get()
    ws["B"+str(row)].value = Month_entry.get()
    ws["C"+str(row)].value = Bama_entry.get()
    ws["D"+str(row)].value = Luong_entry.get()
    ws["E"+str(row)].value = an_entry.get()
    ws["F"+str(row)].value = nha_entry.get()
    ws["G"+str(row)].value = dien_entry.get()
    ws["H"+str(row)].value = nuoc_entry.get()
    ws["I"+str(row)].value = xang_entry.get()
    ws["J"+str(row)].value = dt_entry.get()
    ws["K"+str(row)].value = vdsh_entry.get()
    ws["L"+str(row)].value = toeic_entry.get()
    ws["M"+str(row)].value = thuoc_entry.get()
    ws["N"+str(row)].value = khac_entry.get()
    ws["O"+str(row)].value = chi_entry.get()
    ws["P"+str(row)].value = thu_entry.get()
    ws["Q"+str(row)].value = tiet_kiem_entry.get()
    ws["R"+str(row)].value = Ghi_chu.get("1.0", END)
    ws["S"+str(row)].value = Ghi_chu1.get("1.0", END)
    ws["T"+str(row)].value = Ghi_chu2.get("1.0", END)

    wb.save("Money.xlsx")
    main.destroy()


"------------------------------------------------------------------------------------------------------------------------------------------------"
def Toplevel():
    global Stt_entry, Month_entry, Bama_entry, Luong_entry, an_entry, nha_entry, dien_entry, xang_entry, dt_entry, vdsh_entry, toeic_entry, thuoc_entry, nuoc_entry, khac_entry, chi_entry, thu_entry, tiet_kiem_entry, Ghi_chu, Ghi_chu1, Ghi_chu2, chi_chiem_entry, main
main = Tk()
main.title("Time Management")
main.geometry("1300x850+200+5")
main.resizable(0, 0)

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set backfround
img = ImageTk.PhotoImage(file= "Images\Money.jpg")
label_im = Label(main, image= img)
label_im.place(x= 0, y=0)

# set big frame
frame = Frame(main, height= 700, width= 1100 ,bg= "#2d3134")
frame.place(x = 100, y = 80)

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set small frame
frame1 = LabelFrame(frame, text="Thu vào", font= ("", 15), height= 140, width= 1099, highlightthickness= 0, bg= "#2d3134", fg= "light green")
frame2 = LabelFrame(frame, text= "Chi ra", font= ("", 15), height=300, width= 1099, highlightthickness= 0, bg= "#2d3134", fg= "red")
frame3= LabelFrame(frame, text= "Báo cáo", font= ("", 15), height=249, width= 1099, highlightthickness= 0, bg= "#2d3134", fg= "light blue")

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set widget for frame1 ( Thu vào )
Stt = Label(frame1, text= "      Stt :", font= ("", 12), bg= "#2d3134", fg= "light green")
Stt_entry = ttk.Entry(frame1, width= 5, font= (8))

Month = Label(frame1, text= "Month :", font= ("", 12), bg= "#2d3134", fg= "light green")
Month_entry = ttk.Entry(frame1, width= 8, font= (8))

Bama = Label(frame1, text= "Tiền ba má cho :", font= ("", 12), bg= "#2d3134", fg= "light green")
Bama_entry = ttk.Entry(frame1, width= 15, font= (9))

Luong = Label(frame1, text= "        Tiền lương :", font= ("", 12), bg= "#2d3134", fg= "light green")
Luong_entry = ttk.Entry(frame1, width= 15, font= (9))

text = "Ghi chu o đay"
Ghi_chu_label = Label(frame1, text= "Ghi chú :", font= ("", 12), bg= "#2d3134", fg= "#faf9c3")
Ghi_chu = Text(frame1, height=3, width= 50)
Ghi_chu.insert(END, text)

update = Button(frame1, text= "Update", width= 55, height= 1, bg= "#2d3134", fg="light green" , font=("Arial", 10, "bold"), command= update_thu)
delete = Button(frame1, text= "Delete", width= 55, height= 1, bg= "#2d3134", fg="red" , font=("Arial", 10, "bold"), command= delete_thu)

# set position for widget 
frame1.place(x= 0, y= 0)
frame2.place(x= 0, y= 145)
frame3.place(x= 0, y= 450)

Stt.place(x= 5, y= 0)
Month.place(x= 5, y= 30)
Bama.place(x= 250, y= 0)
Luong.place(x= 250, y= 30)

Stt_entry.place(x= 60, y= 0)
Month_entry.place(x = 60, y= 30)
Bama_entry.place(x= 370, y= 0)
Luong_entry.place(x= 370, y= 30)

Ghi_chu_label.place(x= 585, y= 0)
Ghi_chu.place(x= 650, y= 0)

update.place(x= 70, y= 70)
delete.place(x= 550, y= 70)
# end widget frame 1 ( Thu vào)

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set widget for frame2 ( Chi ra )

tien_an = Label(frame2, text= "Tiền ăn :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_nha = Label(frame2, text= "Tiền nhà :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_dien = Label(frame2, text= "Tiền điện :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_nuoc = Label(frame2, text= "Tiền nước :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_xang = Label(frame2, text= "Tiền xăng :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_dt = Label(frame2, text= "Tiền điện thoại :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_vdsh = Label(frame2, text= "Vật dụng sinh hoạt :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_toeic = Label(frame2, text= "Học Toeic :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_thuoc = Label(frame2, text= "Thuốc men :", font= ("", 12), bg= "#2d3134", fg= "red")
tien_khac = Label(frame2, text= "Phát sinh khác :", font= ("", 12), bg= "#2d3134", fg= "red")

Ghi_chu_label1 = Label(frame2, text= "Ghi chú :", font= ("", 12), bg= "#2d3134", fg= "#faf9c3")
Ghi_chu1 = Text(frame2, height=8, width= 40)
Ghi_chu1.insert(END, text)

an_entry = ttk.Entry(frame2, width= 15, font= (8))
nha_entry = ttk.Entry(frame2, width= 15, font= (8))
dien_entry = ttk.Entry(frame2, width= 15, font= (8))
nuoc_entry = ttk.Entry(frame2, width= 15, font= (8))
xang_entry = ttk.Entry(frame2, width= 15, font= (8))
dt_entry = ttk.Entry(frame2, width= 15, font= (8))
vdsh_entry = ttk.Entry(frame2, width= 15, font= (8))
toeic_entry = ttk.Entry(frame2, width= 15, font= (8))
thuoc_entry = ttk.Entry(frame2, width= 15, font= (8))
khac_entry = ttk.Entry(frame2, width= 15, font= (8))

update1 = Button(frame2, text= "Update", width= 40, height= 2, bg= "#2d3134", fg="light green" , font=("Arial", 10, "bold"), command= update_chi)
delete1 = Button(frame2, text= "Delete", width= 40, height= 2, bg= "#2d3134", fg="red" , font=("Arial", 10, "bold"), command= lambda: delete_chi("1"))

# set position for widget 
tien_an.place(x= 0, y = 5)
tien_nha.place(x= 0, y = 55)
tien_dien.place(x= 0,y = 105)
tien_nuoc.place(x= 0,y = 155)
tien_xang.place(x= 0,y = 205)
tien_dt.place(x= 350, y= 5)
tien_vdsh.place(x= 350, y= 55)
tien_toeic.place(x= 350, y= 105)
tien_thuoc.place(x= 350, y= 155)
tien_khac.place(x= 350, y= 205)

an_entry.place(x= 83, y= 5)
nha_entry.place(x= 83, y= 55)
dien_entry.place(x= 83, y= 105)
nuoc_entry.place(x= 83, y= 155)
xang_entry.place(x= 83, y= 205)
dt_entry.place(x= 500, y= 5)
vdsh_entry.place(x= 500, y= 55)
toeic_entry.place(x= 500, y= 105)
thuoc_entry.place(x= 500, y= 155)
khac_entry.place(x= 500, y= 205)

Ghi_chu_label1.place(x= 680, y= 5)
Ghi_chu1.place(x= 750, y= 5)

update1.place(x= 730, y= 150)
delete1.place(x= 730, y= 199)
# end widget for frame 2 (Chi ra)

"------------------------------------------------------------------------------------------------------------------------------------------------"

# set widget for frame 3 ( Báo cáo )
tong_thu = Label(frame3, text= "Tổng thu :",  font= ("", 12), bg= "#2d3134", fg= "light blue")
tong_chi = Label(frame3, text= "Tổng chi :", font= ("", 12), bg= "#2d3134", fg= "light blue")
chi_chiem = Label(frame3, text= "Chi chiếm :", font= ("", 12), bg= "#2d3134", fg= "light blue")
tiet_kiem = Label(frame3, text= "Tiết kiệm được :", font= ("", 12), bg= "#2d3134", fg= "light blue")

thu_entry = ttk.Entry(frame3, width= 15, font= (8), state= "readonly")
chi_entry = ttk.Entry(frame3, width= 15, font= (8), state= "readonly")
chi_chiem_entry = ttk.Entry(frame3, width= 15, font= (8), state= "readonly")
tiet_kiem_entry = ttk.Entry(frame3, width= 15, font= (8), state= "readonly")

Ghi_chu_label2 = Label(frame3, text= "Ghi chú :", font= ("", 12), bg= "#2d3134", fg= "#faf9c3")
Ghi_chu2 = Text(frame3, height=8, width= 40)
Ghi_chu2.insert(END, text)

update2 = Button(frame3, text= "Update", width= 13, height=4, bg= "#2d3134", fg="light green" , font=("Arial", 10, "bold"), command= updateBaoCao)
save = Button(frame3, text= "Save", width= 40, height= 2, bg= "#2d3134", fg="light blue" , font=("Arial", 10, "bold"), command= Save)

# set position widget
tong_thu.place(x= 5, y = 5)
tong_chi.place(x= 5, y= 55)
chi_chiem.place(x= 300, y= 5)
tiet_kiem.place(x= 300, y= 55)

thu_entry.place(x= 100, y= 5)
chi_entry.place(x= 100, y=55)
chi_chiem_entry.place(x= 445, y= 5)
tiet_kiem_entry.place(x= 445, y= 55)

Ghi_chu_label2.place(x= 650,y = 5)
Ghi_chu2.place(x= 750, y= 5)

update2.place(x= 610, y= 45)
save.place(x= 200, y= 130)

# end widget for frame 3 ( Báo cáo )

"------------------------------------------------------------------------------------------------------------------------------------------------"



mainloop()